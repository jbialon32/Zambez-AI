'''
Created on Jan 17, 2018

Made for Zambez-AI

@author: Kalerne aka dBasshunterb
'''
from openpyxl import load_workbook
from fuzzywuzzy import fuzz


def Monster(bot):
    @bot.command(pass_context=True)
    async def monster(ctx, monsterName=None):
        
        monsterName = monsterName.replace('_', ' ')
        monsterName = monsterName.capitalize()
        monsterBook = load_workbook(filename='Aris-5e-Monster-Sorter.xlsx')
        monsterSheet = monsterBook['Sheet1']
        colA = monsterSheet['A']
        colInfo = colA[1:450]
        monNames = []
        possibleMatches = []
        
        for info in colInfo:
            monName = info.value
            monNames.append(monName)
        
        for name in monNames:
            fuzzMatch = int(fuzz.ratio(monsterName, name))
            if fuzzMatch >= 50:
                fuzzDict = [fuzzMatch, name]
                possibleMatches.append(fuzzDict)
        
        loopCycle = 0
        while loopCycle <= 3:
            for thingy in possibleMatches[loopCycle:]:
                if thingy[0] > possibleMatches[loopCycle][0]:
                    possibleMatches.remove(thingy)
                    possibleMatches.insert(loopCycle, thingy)
            loopCycle += 1
        
        if monsterName not in monNames:
                if len(possibleMatches) > 4:
                    await bot.say("Did you mean... ".format(monsterName))
                    matchLength = 4
                    listPos = 0
                    while matchLength > 0:
                        await bot.say("{}\n".format(possibleMatches[listPos][1]))
                        listPos += 1
                        matchLength -= 1
                    return
                        
                if len(possibleMatches) < 5 and len(possibleMatches) > 0:
                    await bot.say("Did you mean... ".format(monsterName))
                    for match in possibleMatches:
                        await bot.say("{}\n".format(match[1]))
                    return
                            
                if len(possibleMatches) == 0:        
                    await bot.say("{} does not exist".format(monsterName))
                    return
            
        for name in monNames:
            if name == str(monsterName):
                for row in monsterSheet.rows:
                    for cell in row:
                        if cell.value == name:
                            rowVal = cell.row
                            break
                monsterName = name
                monsterCr = monsterSheet.cell(row=rowVal, column=2).value
                monsterType = monsterSheet.cell(row=rowVal, column=3).value
                monsterSubType = monsterSheet.cell(row=rowVal, column=4).value
                monsterSize = monsterSheet.cell(row=rowVal, column=5).value
                monsterAlign = monsterSheet.cell(row=rowVal, column=6).value
                monsterLegendary = monsterSheet.cell(row=rowVal, column=7)
                monsterLair = monsterSheet.cell(row=rowVal, column=8)
                    
                if monsterLegendary.value == "N":
                    monsterLegendary = 'No'
                    
                elif monsterLegendary.value == "Y":
                    monsterLegendary = 'Yes'
                    
                if monsterLair.value == "N":
                    monsterLair = 'No'
                        
                elif monsterLair.value == "Y":
                    monsterLair = 'Yes'
                        
                elif monsterLair.value == "Y (regional only)":
                    monsterLair = 'Yes (Regional Only)'
                    
                await bot.say("Monster Name: {}\n".format(name) +
                                "Monster CR: {}\n".format(monsterCr) +
                                "Monster Type: {}\n".format(monsterType) +
                                "Monster SubType: {}\n".format(monsterSubType) +
                                "Monster Size: {}\n".format(monsterSize) +
                                "Monster Alignment: {}\n".format(monsterAlign) +
                                "Legendary: {}\n".format(monsterLegendary) +
                                "Lair: {}\n".format(monsterLair))
                return
                    