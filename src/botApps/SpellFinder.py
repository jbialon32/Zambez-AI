'''
Created on Jan 18, 2018

Made for Zambez-AI

@author: Kalerne aka dBasshunterb
'''

import discord
from openpyxl import load_workbook
from fuzzywuzzy import fuzz
from botApps.Dice import proDice

    
def SpellFinder(bot=None):
        
    @bot.command(pass_context=True)
    async def spell(ctx, spellName=None, action=None, target: discord.Member=None):
        spellName = spellName.replace('_', ' ')
        spellName = spellName.title()
        """
        
        --------------------------------------------------------
        ONLY TO BE USED AS A COMPANION TO THE OFFICIAL 5e BOOKS!
        --------------------------------------------------------
        
        #################
        BIG THANKS TOO
        Zolo49
        #################
        
        ================================================================================
        https://www.reddit.com/r/DnD/comments/2qs89e/5e_spell_reference_sheets_are_done/
        ================================================================================
        
        """
        spellBook = load_workbook(filename='5E Spells.xlsx')
        spellSheet = spellBook['Master Table']
        colA = spellSheet['A']
        colInfo = colA[5:366]
        spellNames = []
        spellClasses = []
        possibleMatches = []
        
        for info in colInfo:
            spellNameHolder = info.value
            spellNames.append(spellNameHolder)
        
        for name in spellNames:
            fuzzMatch = int(fuzz.ratio(spellName, name))
            if fuzzMatch >= 50:
                fuzzDict = [fuzzMatch, name]
                possibleMatches.append(fuzzDict)
        
        loopCycle = 0
        while loopCycle <= 3:
            for thingy in possibleMatches[loopCycle:]:
                if thingy[0] > possibleMatches[loopCycle][0]:
                    possibleMatches.remove(thingy)
                    possibleMatches.insert(loopCycle, thingy)
            loopCycle += 1
        
        if spellName not in spellNames:
                if len(possibleMatches) > 4:
                    await bot.say("Did you mean... ".format(spellName))
                    matchLength = 4
                    listPos = 0
                    while matchLength > 0:
                        await bot.say("{}\n".format(possibleMatches[listPos][1]))
                        listPos += 1
                        matchLength -= 1
                    return
                        
                if len(possibleMatches) < 5 and len(possibleMatches) > 0:
                    await bot.say("Did you mean... ".format(spellName))
                    for match in possibleMatches:
                        await bot.say("{}\n".format(match[1]))
                    return
                            
                if len(possibleMatches) == 0:        
                    await bot.say("{} does not exist".format(spellName))
                    return
            
        for name in spellNames:
            if name == str(spellName):
                for row in spellSheet.rows:
                    for cell in row:
                        if cell.value == name:
                            rowVal = cell.row
                            break
                spellName = name
                spellLevel= spellSheet.cell(row=rowVal, column=2).value
                spellSchool = spellSheet.cell(row=rowVal, column=3).value
                spellRitual= spellSheet.cell(row=rowVal, column=4).value
                castTime = spellSheet.cell(row=rowVal, column=5).value
                spellRange = spellSheet.cell(row=rowVal, column=6).value
                spellTargets = spellSheet.cell(row=rowVal, column=7).value
                spellVerbal = spellSheet.cell(row=rowVal, column=8).value
                spellSomatic = spellSheet.cell(row=rowVal, column=9).value
                spellMaterial = spellSheet.cell(row=rowVal, column=10).value
                spellComponents = spellSheet.cell(row=rowVal, column=11).value
                spellCost = spellSheet.cell(row=rowVal, column=12).value
                spellConcentration = spellSheet.cell(row=rowVal, column=13).value
                spellDuration = spellSheet.cell(row=rowVal, column=14).value
                spellAtkSave = spellSheet.cell(row=rowVal, column=15).value
                spellDamageType = spellSheet.cell(row=rowVal, column=16).value
                spellDmgHeal = spellSheet.cell(row=rowVal, column=17).value
                spellSourceBook = spellSheet.cell(row=rowVal, column=18).value
                spellSourcePage = spellSheet.cell(row=rowVal, column=19).value
                spellDetail = spellSheet.cell(row=rowVal, column=20).value
                spellLevelBoost = spellSheet.cell(row=rowVal, column=21).value
                
                loopColumn = 22
                while loopColumn <= 29:
                    spellClassCheck = spellSheet.cell(row=rowVal, column=loopColumn).value
                    if spellClassCheck != None:
                        spellClasses.append(spellClassCheck)
                    loopColumn += 1
            
                if spellVerbal == "V":
                    spellVerbal = 'Yes'
                    
                if spellSomatic == "S":
                    spellSomatic = 'Yes'
                    
                if spellMaterial == "M":
                    spellMaterial = 'Yes'
                    
                if spellRitual == "Ritual":
                    spellRitual = 'Yes'
                        
                if spellVerbal == None:
                    spellVerbal = 'No'
                    
                if spellSomatic == None:
                    spellSomatic = 'No'
                    
                if spellMaterial == None:
                    spellMaterial = 'No'
                
                action = action.title()
                
                
                if action == 'Cast':
                    try:
                        diceInfo = spellDmgHeal.split()
                        diceInfo = str(diceInfo[1])
                        betterInfo = []
                        for char in diceInfo:
                            betterInfo.append(char)
                        numDice = int(betterInfo[0])
                        if len(diceInfo) > 3:
                            numSides = str(betterInfo[2]) + str(betterInfo[3])
                            numSides = int(betterInfo[2]) + int(betterInfo[3])
                        else:
                            numSides = str(betterInfo[2])
                            numSides = int(betterInfo[2]) 
                        roll = proDice(numSides, numDice)
                        await bot.say("{} casts {} on {} for {} damage!".format(ctx.message.author.mention, name, target.mention, roll))
                
                    except:
                        await bot.say("Please use proper syntax.\n" +
                                      'spell [name] cast [user (ex: billbob#1244)]\n' +
                                      '\nOr use spell [name] to find a spell\n' +
                                      'and make sure it deals damage.'
                                      )
                    
                
                if action == None:
                        
                    await bot.say("Spell Name:    {}\n".format(name) +
                                    "Level Req:    {}\n".format(spellLevel) +
                                    "School:    {}\n".format(spellSchool) +
                                    "Ritual:    {}\n".format(spellRitual) +
                                    "Cast Time:    {}\n".format(castTime) +
                                    "Range:    {}\n".format(spellRange) +
                                    "Target/Area:    {}\n".format(spellTargets) +
                                    "Verbal:    {}\n".format(spellVerbal) +
                                    "Somatic:    {}\n".format(spellSomatic) +
                                    "Material:    {}\n".format(spellMaterial) +
                                    "Components:    {}\n".format(spellComponents) +
                                    "Cost:    {}\n".format(spellCost) +
                                    "Concentration:    {}\n".format(spellConcentration) +
                                    "Duration:    {}\n".format(spellDuration) +
                                    "Attack/Saving Throw:    {}\n".format(spellAtkSave) +
                                    "Damage Type:    {}\n".format(spellDamageType) +
                                    "Dmg/Heal:    {}\n".format(spellDmgHeal) +
                                    "Source Book:    {}\n".format(spellSourceBook) +
                                    "Source Page:    {}\n".format(spellSourcePage) +
                                    "Detail:    {}\n".format(spellDetail) +
                                    "Level Bonus:    {}\n".format(spellLevelBoost) +
                                    "Classes:    {}".format((', ').join(spellClasses))
                                    )
                                    
                    return

    
                    
